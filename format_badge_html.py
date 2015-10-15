# https://openpyxl.readthedocs.org/en/latest/
# cd /Users/jake.spirek/GitHub/badges 
# python3 format_badge_html.py 

import webbrowser
import os
os.system("clear")
print("\n ----- Begin Script ----- \n")


# Set up spreadsheet
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter, column_index_from_string

# Load spreadsheet
wb = load_workbook('badge_descriptions.xlsx', guess_types=True)

# Grab the active worksheet by name
ws = wb['Working']
num_rows = len(ws.rows)
num_cols = len(ws.columns)
id_list = []
counter = 1

def find_column_index_by_title(title):
	for cell in ws.rows[0]:
		if cell.value == title:
			# print("Found Ticket Subject in: ", column_index_from_string(cell.column), cell.row, cell.coordinate)
			return column_index_from_string(cell.column)-1


def findColumnByTitle(title):
	for cell in ws.rows[0]:
		if cell.value == title:
			# print("Found Ticket Subject in: ", column_index_from_string(cell.column), cell.row, cell.coordinate)
			return cell.column


def write_to_textfile(thelist):
	# http://www.pythonforbeginners.com/files/reading-and-writing-files-in-python
	thelist.pop(0)
	filepath = "badges.html"
	# profile_url = "http://static.smallworldlabs.com/blackbaud/content/icons/badges/profile/"
	# persistent_url = "http://static.smallworldlabs.com/blackbaud/content/icons/badges/persistent/"
	with open(filepath, 'w') as file:
		file.write('<ul class="media-list">\n')
		for index, item in enumerate(thelist):

			file.write('\t<li class="media">\n\t\t<div class="media-left">\n\t\t\t<img class="media-object" src="http://static.smallworldlabs.com/blackbaud/content/icons/badges/{}/{}" alt="{}">\n\t\t</div>\n'.format(item[badge_type_col].value.lower(), item[badge_filename_col].value, index))
			file.write('\t\t<div class="media-body">\n\t\t\t<h4 class="media-heading">{}</h4>\n\t\t<p>{}</p></div>\n'.format(item[badge_heading_col].value, item[badge_description_col].value))
			file.write('\t</li>\n')
		file.write('</ul>\n')


def cycle_thru_each_row():
	for row in ws.rows:
		# print(row[badge_filename_col-1].value, row[badge_type_col-1].value, row[badge_description_col-1].value)
		# id_list.append(row[badge_filename_col-1].value)
		id_list.append(row)




badge_filename_col = find_column_index_by_title("Filename")
badge_type_col = find_column_index_by_title("Type")
badge_heading_col = find_column_index_by_title("Heading")
badge_description_col = find_column_index_by_title("Description")

print("\n", "Filename Column Index:", badge_filename_col)

# print("\n", "Column A:", ws.columns[badge_filename_col-1])
# print("\n", "Column B:", ws.columns[badge_type_col-1])

# print("\n", "Row 1:", ws.rows[badge_filename_col][0].value)

cycle_thru_each_row()

# test write
write_to_textfile(id_list)


print("this list has {} items".format(len(id_list)))


